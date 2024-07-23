import openpyxl
import smtplib
import os
import sys
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw

# Add participant name to certificate
def make_certi(name):

    img = Image.open("razzsecuritycertificate.png")
    draw = ImageDraw.Draw(img)

    # Load font
    font = ImageFont.truetype("InknutAntiqua-ExtraBold.ttf", 33)

    if name == -1:
        return -1
    else:
        # Insert text into image template
        draw.text( (287,207), name, (0,0,0), font=font )

        if not os.path.exists( 'online_intern_6_jan_2021' ) :
            os.makedirs( 'online_intern_6_jan_2021' )

        # Save as a PDF
        rgb = Image.new('RGB', img.size, (255, 255, 255))  # white background
        rgb.paste(img, mask=img.split()[3])               # paste using alpha channel as mask
        
        rgb.save( 'online_intern_6_jan_2021/'+'RS-cert-'+str(name)+'.pdf', "PDF", resolution=100.0)
        return 'online_intern_6_jan_2021/'+'RS-cert-'+str(name)+'.pdf'

# Email the certificate as an attachment
def email_certi( filename, receiver, name ):
    username = "razzsec.contact@gmail.com"
    password = "hhpiwyfblgdllxxm" 
    sender = username + '@gmail.com'

    msg = MIMEMultipart()
    msg['Subject'] = 'RAZZ SECURITY ETHICAL HACKING AND PENETRATION TESTING COMPLETION CERTIFICATE'
    msg['From'] = username+'@gmail.com'
    msg['Reply-to'] = username + '@gmail.com'
    msg['To'] = receiver

    # That is what u see if dont have an email reader:
    msg.preamble = 'Multipart massage.\n'
    
    # Body
    body = "Congratulations! {}\nYou have successfully participated in webinar on Roadmap to cyber security. \nWe hope that this course fulfilled your expectations !\nThank you & Regards\nRazz Security Team".format(name.upper())
    part = MIMEText(body)
    msg.attach( part )

    # Attachment
    part = MIMEApplication(open(filename,"rb").read())
    part.add_header('Content-Disposition', 'attachment', filename = os.path.basename(filename))
    msg.attach( part )

    # Login
    server = smtplib.SMTP( 'smtp.gmail.com:587' )
    server.starttls()
    server.login( username, password )

    # Send the email
    server.sendmail( msg['From'], msg['To'], msg.as_string() )

if __name__ == "__main__":
    error_list = []
    error_count = 0

    os.chdir(os.path.dirname(os.path.abspath((sys.argv[0]))))

    # Read data from an excel sheet from row 2
    Book = openpyxl.load_workbook('akk.xlsx')
    WorkSheet = Book.active
    
    num_row = WorkSheet.max_row
    row = 0

    for row_num in range(2, num_row + 1):
        name = WorkSheet.cell(row=row_num, column=3).value
        email = WorkSheet.cell(row=row_num, column=2).value
        
        # Make certificate and check if it was successful
        filename = make_certi(name.upper())
        
        # Successfully made certificate
        if filename != -1:
            email_certi(filename, email, name)
            print("Sent to ", name)
        # Add to error list
        else:
            error_list.append(name)  # Append name instead of ID
            error_count += 1

    # Print all failed names
    print(str(error_count), " Errors- List:", ','.join(error_list))
